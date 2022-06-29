#Ayush Sahu - 1901CB13
#Aditya Goyal - 1901EE06
from flask import Flask
from flask import render_template, url_for, request, redirect
import shutil
from os import name
import os.path
app = Flask(__name__)
import pandas as pd
import dataframe_image as dfi
import csv
from openpyxl import Workbook
from openpyxl import load_workbook
import numpy as np
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from tabulate import tabulate
from fpdf import FPDF
from datetime import date
from datetime import datetime
# ndir, master_flag , response_flag, pos, neg,ldir,btn1,btn2,btn3,notlast,cantsend,gen_pos,gen_neg = (0,)*13
path1 = "sample_input"
dir1 = os.path.join(path1)
if os.path.isdir(dir1) == True:
    shutil.rmtree(dir1)
os.mkdir(path1)
folder = "transcriptsIITP"
dir2 = os.path.join(folder)
if os.path.isdir(dir2) == True:
    shutil.rmtree(dir2)
path3 = "Seal_Sign_image"
dir3 = os.path.join(path3)
if os.path.isdir(dir3) == True:
    shutil.rmtree(dir3)
os.mkdir(path3)
loc_csv = os.path.join(app.root_path,path1)   
loc_image = os.path.join(app.root_path,path3)
# print(loc) 
app.config["input_file_loc"] = loc_csv
app.config["image_loc"] = loc_image
names_flag,subjects_flag,grades_flag,Seal_flag,Sign_flag,csv_count_flag,ldir,ndir,submit_1, starts, ends, csv_dir, submit_1_flag, submit_2,submit_2_flag,left_roll_no= (0,)*16
Master_dict = {}
namedict = {}
courses={
    'CS':"Computer Science and Engineering",
    'EE':"Electrical Engineering",
    'ME':"Mechanical Engineering",
    'CB':"Chemical and Biochemical Engineering",
    'CE':"Civil Engineering",
    "MM":"Metallurgical Engineering"
}
Seal=""
Sign=""

def generate_marksheet():
    #If output folder is present then removing it and then making it again for avoiding readding of data
    #Mde four dictionaries for storing the data from names-roll.csv and subjects_master.csv
    subnames = {}
    subltp = {}
    Discipline = {}
    #storing data from names-roll.csv
    with open(os.path.join('sample_input','names-roll.csv')) as file:
        namesroll = csv.DictReader(file)
        for i in namesroll:
            namedict[i['Roll'].strip()] = i['Name']
            D = ""
            for char in i['Roll']:
                D += char if ((ord(char)>=65 and ord(char)<=90) or (ord(char) >= 97 and ord(char) <= 122)) else ""
            Discipline[i['Roll'].strip()] = D;
    #storing data from subjects_master.csv          
    with open(os.path.join('sample_input','subjects_master.csv')) as file:
        subjects = csv.DictReader(file)
        for i in subjects:
            subnames[i['subno']] = i['subname']
            subltp[i['subno']] = i['ltp']
    #Now iterating on grades.csv
    with open(os.path.join('sample_input','grades.csv')) as file:
        grades = csv.DictReader(file)
        for i in grades:
            #Checking if the roll_no.xlsx file is present in output folder or not
            roll = i['Roll'].strip()

            if roll in Master_dict.keys():
                #As the file is present, so laoding the workbook
                Sem = i['Sem']
                sem_no = "Sem"+Sem
                #Checking if a sheet with Sem# is present or not.
                if sem_no in Master_dict[roll].keys():
                    #As is it is present, so just appending the data from grade.csv in the sheet.
                    Master_dict[roll][sem_no].append([i['SubCode'].strip(),subnames[i['SubCode'].strip()],subltp[i['SubCode'].strip()],i['Credit'],i['Grade'].strip()])
                else:
                    #As it is not present, so creating a new Sheet in the required roll_no.xlsx file and adding the initial contents
                    Master_dict[roll][sem_no] = []
                    Master_dict[roll][sem_no].append(['Sub. Code','Subject Name','L-T-P','CRD','GRD'])
                    si = 1
                    Master_dict[roll][sem_no].append([i['SubCode'].strip(),subnames[i['SubCode'].strip()],subltp[i['SubCode'].strip()],i['Credit'],i['Grade'].strip()])
            else:
                #As the file is not present, so making a new Workbook
                Sem = i['Sem']
                sem_no = "Sem"+Sem
                Master_dict[roll] = {}
                Master_dict[roll]["Overall"] = []
                Master_dict[roll][sem_no] = []
                Master_dict[roll][sem_no].append(['Sub. Code','Subject Name','L-T-P','CRD','GRD'])
                Master_dict[roll][sem_no].append([i['SubCode'].strip(),subnames[i['SubCode'].strip()],subltp[i['SubCode'].strip()],i['Credit'],i['Grade'].strip()])
    

    #Now in every roll_no file present, we have made every sem sheet and added its content, but we have not added content in overall sheet, so iterating over names-roll.csv for going roll no wise to add overall sheet contents
    with open(os.path.join('sample_input','names-roll.csv')) as file:
	#A dictionary for taking care of grades 
        gradedict = {'AA':10,'AB':9,'BB':8,'BC':7,'CC':6,'CD':5,'DD':4,'F':0,'I':0,'DD*':4,'F*':0}
        finalprf = csv.DictReader(file)
        global sem_credits_clear
        for f in finalprf:
            flag = 0
            #storing file path of the roll_no.xlsx in target variable
            roll = f['Roll'].strip()
            if roll in Master_dict.keys():
                #Firsty adding simple contents like Roll no, name of student and Discipline
                Master_dict[roll]["Overall"].append(['Roll No.',f['Roll'].strip()])
                Master_dict[roll]["Overall"].append(['Name of Student',namedict[f['Roll'].strip()]])
                Master_dict[roll]["Overall"].append(['Discipline',Discipline[f['Roll'].strip()]])
                #Now making lists of other data to be appended after calculation.
                semlist = ['Semester No.']
                semcredit = ['Semester Wise Credit Taken']
                totalcredits = ['Total Credits Taken']
                spi = ['SPI']
                cpi = ['CPI']
                sem_credits_clear = ['Credits Clear']
                creditssum = 0
                #Now we will be iterating every sem sheet of the roll_no we are currently on.
                for n in Master_dict[roll].keys():
                    #Now this condition is for avoiding interence with Overall sheet, we will be adding data in Overall sheet, but firstly we need to collect the data.
                    if flag==1:
                        #num is making extracting sem no.
                        num = n[3:]
                        semlist.append(num)
                        sem_list = Master_dict[roll][n]
                        creditpersem = 0
                        credits_clear = 0
                        credits_fail = 0
                        credit = []
                        gradepersubject = []
                        #Now iterating in the sheet we are currently on.
                        for i in range(0,len(sem_list)):
                            for j in range(0,len(sem_list[i])):
                                cell_obj = sem_list[i][j]
                                if(i>0 and j==3):
                                    #Credit column data
                                    creditpersem+=int(cell_obj)
                                    credit.append(int(cell_obj))
                                if(i>0 and j==4):
                                    #Grade column data
                                    if cell_obj=="F" or cell_obj == "F*":
                                        credits_fail += int(sem_list[i][j-1])
                                    gradepersubject.append(cell_obj)
                        sum = 0           
                        for k in range(0,len(credit)):
                            sum += credit[k]*gradedict[gradepersubject[k]]  
                        #spi calculation
                        spisem = round(sum/creditpersem,2)
                        #adding credits after every sem
                        creditssum += creditpersem
                        credits_clear = creditpersem-credits_fail
                        #Now appending the data
                        sem_credits_clear.append(credits_clear)
                        semcredit.append(creditpersem)
                        spi.append(spisem)
                        totalcredits.append(creditssum)
                    flag = 1
                cpi.append(spi[1])
                #cpi calculation
                for i in range(1,len(spi)-1):
                    numerator = cpi[i]*totalcredits[i] + spi[i+1]*semcredit[i+1]
                    denominator = totalcredits[i+1]
                    cpipersem = round(numerator/denominator,2)
                    cpi.append(cpipersem)
                #Now adding the collected data in the Overall sheet
                Master_dict[roll]["Overall"].append(semlist)  
                Master_dict[roll]["Overall"].append(semcredit)
                Master_dict[roll]["Overall"].append(spi)
                Master_dict[roll]["Overall"].append(totalcredits)
                Master_dict[roll]["Overall"].append(cpi)
                Master_dict[roll]["Overall"].append(sem_credits_clear)
          
class PDF(FPDF):
     def lines(self):
            self.rect(9, 6.5, 400.0,273.0)
            self.rect(90, 46, 240.0,15.0)
            print(Seal,Sign)
     def imagex(self):
         self.image("templates/heading.jpg", x = 9.4, y = 6.7, w = 400, h = 38)
         if Seal!="":
           self.image("Seal_Sign_image/Seal.jpeg", x = 160, y = 230, w = 43, h = 43)
         if Sign!="":
           self.image("Seal_Sign_image/Sign.jpeg", x = 330, y = 225, w = 40, h = 40)
                  
def generate_transcript_range(l,r):
    if os.path.exists(folder)==True:
        shutil.rmtree(folder)  
    os.mkdir(folder)
    l=l.upper()
    r=r.upper()
    code=l[0:6]
    number=int(l[6:])
    code1=r[0:6]
    if code!=code1:
        raise ValueError
        return
    left=[]
    flag=0
    while flag==0:
        f=l+".pdf"
        path_file=os.path.join(folder,f)
        if l==r:
            flag=1
        if l in Master_dict:
            pdf=PDF(format='A3',orientation='landscape')
            pdf.add_page()
            pdf.lines()
            pdf.imagex()
            pdf.set_xy(91,45)
            pdf.set_font('helvetica', 'B', 12.0)
            pdf.cell(10,10,"Roll No:",border=0,align='L')
            pdf.set_xy(91,51)
            pdf.set_font('helvetica', 'B', 12.0)
            pdf.cell(10,10,"Programme:",border=0,align='L')
            pdf.set_xy(110,45)
            pdf.set_font('helvetica', '', 12.0)
            pdf.cell(10,10,l,border=0,align='L')
            pdf.set_xy(117,51)
            pdf.set_font('helvetica', '', 12.0)
            pdf.cell(10,10,"Bachelor of Technology",border=0,align='L')
            pdf.set_xy(170,45)
            pdf.set_font('helvetica', 'B', 12.0)
            pdf.cell(10,10,"Name:",border=0,align='L')
            pdf.set_xy(170,51)
            pdf.set_font('helvetica', 'B', 12.0)
            pdf.cell(10,10,"Course:",border=0,align='L')
            pdf.set_xy(186,45)
            pdf.set_font('helvetica', '', 12.0)
            pdf.cell(10,10,namedict[l],border=0,align='L')
            pdf.set_xy(270,45)
            pdf.set_font('helvetica', 'B', 12.0)
            pdf.cell(10,10,"Year of Admission:",border=0,align='L')
            pdf.set_xy(310,45)
            pdf.set_font('helvetica', '', 12.0)
            pdf.cell(10,10,"20"+code[0:2],border=0,align='L')
            pdf.set_xy(188,51)
            pdf.set_font('helvetica', '', 12.0)
            pdf.cell(10,10,courses[code[4:6]],border=0,align='L')
            pdf.set_xy(13,255)
            pdf.set_font('helvetica', 'B', 16.0)
            today = date.today()
            now = datetime.now()
            pdf.cell(10,10,"Date Generated: "+today.strftime("%d %b %Y")+", "+now.strftime("%H:%M"),border=0,align='L')
            pdf.set_xy(330,255)
            pdf.set_font('helvetica', 'B', 14.0)
            pdf.cell(10,10,"Assitant Registrar (Academic):",border=0,align='L')
            epw=pdf.w-2*pdf.l_margin
            sem_width=epw/3.3
            col_width=[0.15*sem_width,0.65*sem_width,0.10*sem_width,0.075*sem_width,0.075*sem_width]
            max_y=0
            overall_sem_info = {}
            temp_y=0
            total=len(Master_dict[l])
            i1=0
            for key in Master_dict[l]:
                if key=="Overall":
                    df = pd.DataFrame(Master_dict[l][key])
                    particular = Master_dict[l][key]
                    flag_for_skip = 0
                    cnt = 0
                    for i in particular[3]:
                       if flag_for_skip==1:
                         cnt += 1
                         Sem = "Sem"+i
                         overall_sem_info[Sem] = "Credits Taken: "+str(particular[4][cnt])+"   Credits Cleared: "+str(particular[8][cnt])+"   SPI: "+str(particular[5][cnt])+"   CPI: "+str(particular[7][cnt])
                       flag_for_skip = 1 
                else:
                    # sem_no=int(key[3])
                    if i1<=3:
                        pdf.set_xy(12+132*(i1-1),60)
                        pdf.set_font('helvetica', 'BU', 10.0)
                        pdf.cell(10,10,"Semester "+key[3],border=0,align='L')
                        pdf.set_xy(13+132*(i1-1),68)
                        pdf.set_font('helvetica', '', 9.0)
                        for row in Master_dict[l][key]:
                            col=0
                            for data in row:
                               pdf.cell(col_width[col],4,str(data),border=1,align='C')
                               col=col+1
                            pdf.set_xy(13+132*(i1-1),pdf.get_y()+4)
                        # print(pdf.get_y())
                        max_y=max(max_y,pdf.get_y()+4)
                        pdf.set_xy(13+132*(i1-1),pdf.get_y()+2)
                        pdf.set_font('helvetica', 'B', 9.0)
                        pdf.cell(110,6,overall_sem_info[key],border=1,align='L')
                        if i1==3 or i1==total-1:
                          pdf.set_line_width(0.5)
                          pdf.line(9,max_y+6,409,max_y+6)
                    elif i1<=6:
                        if i1==4:
                          temp_y=max_y+6
                        tc=i1-3
                        pdf.set_xy(12+132*(tc-1),temp_y)
                        pdf.set_font('helvetica', 'BU', 10.0)
                        pdf.cell(10,10,"Semester "+key[3],border=0,align='L')
                        pdf.set_xy(13+132*(tc-1),temp_y+8)
                        pdf.set_font('helvetica', '', 9.0)
                        for row in Master_dict[l][key]:
                            col=0
                            for data in row:
                               pdf.cell(col_width[col],4,str(data),border=1,align='C')
                               col=col+1
                            pdf.set_xy(13+132*(tc-1),pdf.get_y()+4)
                        # print(pdf.get_y())
                        max_y=max(max_y,pdf.get_y()+4)
                        pdf.set_xy(13+132*(tc-1),pdf.get_y()+2)
                        pdf.set_font('helvetica', 'B', 9.0)
                        pdf.cell(110,6,overall_sem_info[key],border=1,align='L')
                        if i1==6 or i1==total-1:
                          pdf.set_line_width(0.5)
                          pdf.line(9,max_y+6,409,max_y+6)
                    else:
                        if i1==7:
                              temp_y=max_y+6
                        tc=i1-6
                        pdf.set_xy(12+132*(tc-1),temp_y)
                        pdf.set_font('helvetica', 'BU', 10.0)
                        pdf.cell(10,10,"Semester "+key[3:],border=0,align='L')
                        pdf.set_xy(13+132*(tc-1),temp_y+8)
                        pdf.set_font('helvetica', '', 9.0)
                        for row in Master_dict[l][key]:
                            col=0
                            for data in row:
                               pdf.cell(col_width[col],4,str(data),border=1,align='C')
                               col=col+1
                            pdf.set_xy(13+132*(tc-1),pdf.get_y()+4)
                        max_y=max(max_y,pdf.get_y()+4)
                        pdf.set_xy(13+132*(tc-1),pdf.get_y()+2)
                        pdf.set_font('helvetica', 'B', 9.0)
                        pdf.cell(110,6,overall_sem_info[key],border=1,align='L')
                        if i1==9 or i1==total-1:
                          pdf.set_line_width(0.5)
                          pdf.line(9,max_y+6,409,max_y+6)
                i1=i1+1 
            pdf.output(path_file,'F') 
        else:
            left.append(l)
        number=number+1
        if number<10:
          l=code+'0'+str(number)
        else:
          l=code+str(number)
    return left
    
def generate_transcripts_all():
    if os.path.exists(folder)==True:
        shutil.rmtree(folder)  
    os.mkdir(folder)
    with open(os.path.join('sample_input','names-roll.csv')) as file:
        names_roll = csv.DictReader(file)
        for line in names_roll:
            roll_no = line["Roll"].strip().upper()
            f=roll_no+".pdf"
            path_file=os.path.join(folder,f)
            
            pdf=PDF(format='A3',orientation='landscape')
            pdf.add_page()
            pdf.lines()
            pdf.imagex()
            pdf.set_xy(91,45)
            pdf.set_font('helvetica', 'B', 12.0)
            pdf.cell(10,10,"Roll No:",border=0,align='L')
            pdf.set_xy(91,51)
            pdf.set_font('helvetica', 'B', 12.0)
            pdf.cell(10,10,"Programme:",border=0,align='L')
            pdf.set_xy(110,45)
            pdf.set_font('helvetica', '', 12.0)
            pdf.cell(10,10,roll_no,border=0,align='L')
            pdf.set_xy(117,51)
            pdf.set_font('helvetica', '', 12.0)
            pdf.cell(10,10,"Bachelor of Technology",border=0,align='L')
            pdf.set_xy(170,45)
            pdf.set_font('helvetica', 'B', 12.0)
            pdf.cell(10,10,"Name:",border=0,align='L')
            pdf.set_xy(170,51)
            pdf.set_font('helvetica', 'B', 12.0)
            pdf.cell(10,10,"Course:",border=0,align='L')
            pdf.set_xy(186,45)
            pdf.set_font('helvetica', '', 12.0)
            pdf.cell(10,10,namedict[roll_no],border=0,align='L')
            pdf.set_xy(270,45)
            pdf.set_font('helvetica', 'B', 12.0)
            pdf.cell(10,10,"Year of Admission:",border=0,align='L')
            pdf.set_xy(310,45)
            pdf.set_font('helvetica', '', 12.0)
            pdf.cell(10,10,"20"+roll_no[0:2],border=0,align='L')
            pdf.set_xy(188,51)
            pdf.set_font('helvetica', '', 12.0)
            pdf.cell(10,10,courses[roll_no[4:6]],border=0,align='L')
            pdf.set_xy(13,255)
            pdf.set_font('helvetica', 'B', 16.0)
            today = date.today()
            now = datetime.now()
            pdf.cell(10,10,"Date Generated: "+today.strftime("%d %b %Y")+", "+now.strftime("%H:%M"),border=0,align='L')
            pdf.set_xy(330,255)
            pdf.set_font('helvetica', 'B', 14.0)
            pdf.cell(10,10,"Assitant Registrar (Academic):",border=0,align='L')
            epw=pdf.w-2*pdf.l_margin
            sem_width=epw/3.3
            col_width=[0.15*sem_width,0.65*sem_width,0.10*sem_width,0.075*sem_width,0.075*sem_width]
            max_y=0
            overall_sem_info = {}
            temp_y=0
            total=len(Master_dict[roll_no])
            i1=0
            for key in Master_dict[roll_no]:
                if key=="Overall":
                    df = pd.DataFrame(Master_dict[roll_no][key])
                    particular = Master_dict[roll_no][key]
                    flag_for_skip = 0
                    cnt = 0
                    for i in particular[3]:
                        if flag_for_skip==1:
                            cnt += 1
                            Sem = "Sem"+i
                            overall_sem_info[Sem] = "Credits Taken: "+str(particular[4][cnt])+"   Credits Cleared: "+str(particular[8][cnt])+"   SPI: "+str(particular[5][cnt])+"   CPI: "+str(particular[7][cnt])
                        flag_for_skip = 1 
                else:
                    # sem_no=int(key[3])
                    if i1<=3:
                        pdf.set_xy(12+132*(i1-1),60)
                        pdf.set_font('helvetica', 'BU', 10.0)
                        pdf.cell(10,10,"Semester "+key[3],border=0,align='L')
                        pdf.set_xy(13+132*(i1-1),68)
                        pdf.set_font('helvetica', '', 9.0)
                        for row in Master_dict[roll_no][key]:
                            col=0
                            for data in row:
                                pdf.cell(col_width[col],4,str(data),border=1,align='C')
                                col=col+1
                            pdf.set_xy(13+132*(i1-1),pdf.get_y()+4)
                        # print(pdf.get_y())
                        max_y=max(max_y,pdf.get_y()+4)
                        pdf.set_xy(13+132*(i1-1),pdf.get_y()+2)
                        pdf.set_font('helvetica', 'B', 9.0)
                        pdf.cell(110,6,overall_sem_info[key],border=1,align='L')
                        if i1==3 or i1==total-1:
                            pdf.set_line_width(0.5)
                            pdf.line(9,max_y+6,409,max_y+6)
                    elif i1<=6:
                        if i1==4:
                            temp_y=max_y+6
                        tc=i1-3
                        pdf.set_xy(12+132*(tc-1),temp_y)
                        pdf.set_font('helvetica', 'BU', 10.0)
                        pdf.cell(10,10,"Semester "+key[3],border=0,align='L')
                        pdf.set_xy(13+132*(tc-1),temp_y+8)
                        pdf.set_font('helvetica', '', 9.0)
                        for row in Master_dict[roll_no][key]:
                            col=0
                            for data in row:
                                pdf.cell(col_width[col],4,str(data),border=1,align='C')
                                col=col+1
                            pdf.set_xy(13+132*(tc-1),pdf.get_y()+4)
                        # print(pdf.get_y())
                        max_y=max(max_y,pdf.get_y()+4)
                        pdf.set_xy(13+132*(tc-1),pdf.get_y()+2)
                        pdf.set_font('helvetica', 'B', 9.0)
                        pdf.cell(110,6,overall_sem_info[key],border=1,align='L')
                        if i1==6 or i1==total-1:
                            pdf.set_line_width(0.5)
                            pdf.line(9,max_y+6,409,max_y+6)
                    else:
                        if i1==7:
                                temp_y=max_y+6
                        tc=i1-6
                        pdf.set_xy(12+132*(tc-1),temp_y)
                        pdf.set_font('helvetica', 'BU', 10.0)
                        pdf.cell(10,10,"Semester "+key[3:],border=0,align='L')
                        pdf.set_xy(13+132*(tc-1),temp_y+8)
                        pdf.set_font('helvetica', '', 9.0)
                        for row in Master_dict[roll_no][key]:
                            col=0
                            for data in row:
                                pdf.cell(col_width[col],4,str(data),border=1,align='C')
                                col=col+1
                            pdf.set_xy(13+132*(tc-1),pdf.get_y()+4)
                        max_y=max(max_y,pdf.get_y()+4)
                        pdf.set_xy(13+132*(tc-1),pdf.get_y()+2)
                        pdf.set_font('helvetica', 'B', 9.0)
                        pdf.cell(110,6,overall_sem_info[key],border=1,align='L')
                        if i1==9 or i1==total-1:
                            pdf.set_line_width(0.5)
                            pdf.line(9,max_y+6,409,max_y+6)
                i1=i1+1 
            pdf.output(path_file,'F') 
    return

@app.route("/", methods=["GET", "POST"])
def GUI():
    global names_flag,subjects_flag,grades_flag,Seal_flag,Sign_flag,csv_count_flag,ldir,ndir,submit_1, starts, ends, csv_dir,submit_1_flag,submit_2,submit_2_flag,left_roll_no
    if request.method == "POST":
        submit_1 = 0
        submit_1_flag = 0
        submit_2 = 0
        submit_2_flag = 0
        left_roll_no  = []
        if request.files:
            if "csv-file" in request.files.keys():
                file = request.files["csv-file"]
                if file.filename == "names-roll.csv":
                    names_flag = 1
                if file.filename == "subjects_master.csv":
                    subjects_flag = 1
                if file.filename == "grades.csv":
                    grades_flag = 1
                file.save(os.path.join(app.config["input_file_loc"], file.filename))
            elif "Seal-image" in request.files.keys():
                global Seal
                file = request.files["Seal-image"]
                Seal_flag = 1
                Seal = os.path.join(app.config["image_loc"], "Seal.jpeg")
                if os.path.isfile(Seal)==True:
                    os.remove(Seal)
                file.save(os.path.join(app.config["image_loc"], file.filename))
                os.rename(os.path.join(app.config["image_loc"], file.filename),Seal)
            elif "Sign-image" in request.files.keys():
                global Sign
                file = request.files["Sign-image"]
                Sign_flag = 1
                Sign = os.path.join(app.config["image_loc"], "Sign.jpeg")
                if os.path.isfile(Sign)==True:
                    os.remove(Sign)
                file.save(os.path.join(app.config["image_loc"], file.filename))
                os.rename(os.path.join(app.config["image_loc"], file.filename),Sign)
            csv_count_flag = 0
            ldir = os.listdir(loc_csv)+os.listdir(loc_image)
            csv_dir = len(os.listdir(loc_csv))
            ndir = len(ldir)
            if ndir>0:
                csv_count_flag = 1
            return redirect(request.url)
        elif request.form:
            global Master_dict
            if request.form["submit"] == "submit-1":
                submit_1 = 1
                starts = request.form["starts"]
                ends = request.form["ends"]
                Master_dict = {}
                try:
                	generate_marksheet()
                	left_roll_no = generate_transcript_range(starts,ends)
                except:
                    submit_1_flag = -1
            if request.form["submit"] == "submit-2":
                submit_2 = 1
                Master_dict = {}
                try:
                    generate_marksheet()
                    generate_transcripts_all()
                except:
                    submit_2_flag = -1

    return render_template('index.html',value1 = names_flag, value2 = subjects_flag, value3 = grades_flag, value4 = Seal_flag, value5 = Sign_flag,value7 = csv_count_flag,value6 = ldir, value8 = submit_1,value9 = csv_dir, value10 = submit_1_flag, value11 = submit_2, value12 = submit_2_flag,value13 = left_roll_no,value14 = starts, value15 = ends)

if __name__ == "__main__":
    app.run(debug=True,port = 5001)
